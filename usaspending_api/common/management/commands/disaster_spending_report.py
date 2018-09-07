import csv
import glob
import itertools
import json
import logging
import numpy as np
import os
import pandas as pd
import shutil
import zipfile

from datetime import datetime
from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook
from time import perf_counter

from usaspending_api.common.helpers.generic_helper import generate_fiscal_year_and_quarter


TAS_XLSX_FILE = "usaspending_api/data/DEFC ABC Pd 6 FY18.xlsx"

logger = logging.getLogger('console')


def dump_to_csv(filepath, data_lines):
    if len(data_lines) == 0:
        return
    if os.path.exists(filepath):
        os.unlink(filepath)

    with open(filepath, "w") as ofile:
        writer = csv.writer(ofile)
        for row in data_lines:
            writer.writerow(row)


def generate_tas_rendering_label(tas_dict):
    """
        Copied from usaspending_api/accounts/models.py and change to use a dict instead of separate parameters
    """
    tas_rendering_label = "-".join(filter(None, (tas_dict["ATA"], tas_dict["AID"])))

    if tas_dict["AvailType Code"] is not None and tas_dict["AvailType Code"] != "":
        tas_rendering_label = "-".join(filter(None, (tas_rendering_label, tas_dict["AvailType Code"])))
    else:
        poa = "\\".join(filter(None, (tas_dict["BPOA"], tas_dict["EPOA"])))
        tas_rendering_label = "-".join(filter(None, (tas_rendering_label, poa)))

    tas_rendering_label = "-".join(filter(None, (tas_rendering_label, tas_dict["Main"], tas_dict["Sub"])))

    return tas_rendering_label


class Command(BaseCommand):
    """
    """

    help = "Generate CSV files for provided TAS to help track disaster spending"
    logger = logging.getLogger("console")

    def add_arguments(self, parser):
        parser.add_argument(
            "-d",
            "--destination",
            default=os.path.dirname(os.path.abspath(__file__)),
            type=str,
            help="Set for a custom location of output files",
        )
        parser.add_argument("-k", "--keep-files", action="store_true", help="If set, don't delete the temp files files")

    def handle(self, *args, **options):
        script_start = perf_counter()
        self.contract_columns = []
        self.assistance_columns = []
        self.first_contract_write = True
        self.first_assistance_write = True
        self.temporary_dir = os.path.dirname(options["destination"] + "/temp_disaster_tas_csv/")
        self.destination = os.path.dirname(options["destination"] + "/OMB_DHS_disaster_report/")
        self.keep_files = options["keep_files"]
        self.verbose = True if options["verbosity"] > 1 else False

        if not os.path.exists(self.temporary_dir):
            os.makedirs(self.temporary_dir)
        if not os.path.exists(self.destination):
            os.makedirs(self.destination)

        tas_dict_list = self.gather_tas_from_file()
        self.query_database(tas_dict_list)
        self.assemble_csv_files()
        self.cleanup_files()

        logger.info("total script runtime: {}s".format(perf_counter() - script_start))

    def gather_tas_from_file(self):
        wb = load_workbook(filename=TAS_XLSX_FILE, read_only=True)
        ws = wb["DEFC"]
        ws.calculate_dimension()
        tas_code_header = ws["A8":"G8"]
        expected_headers = ["ATA", "AID", "BPOA", "EPOA", "AvailType Code", "Main", "Sub"]
        headers = []
        for header in tas_code_header:
            for cell in header:
                headers.append(cell.value.replace("\n", "").strip())
        if expected_headers != headers:
            raise Exception("Headers {} Don't match expected: {}".format(headers, expected_headers))

        tas_code_rows = ws["A9":"G100"]
        # Turn "rows" of "cell" into a list of dictionaries using the headers. I'm sure pandas could have done it too
        tas_dicts = [
            {key: cell.value.strip() if cell.value else None for key, cell in itertools.zip_longest(headers, row)}
            for row in tas_code_rows
        ]

        return tas_dicts

    def query_database(self, tas_dict_list):
        db_start = perf_counter()

        for i, tas in enumerate(tas_dict_list, 1):
            contract_results = self.single_tas_query(tas, "contract")
            filepath = "{}/{}_fpds.csv".format(self.temporary_dir, generate_tas_rendering_label(tas))
            dump_to_csv(filepath, contract_results)

            assistance_results = self.single_tas_query(tas, "assistance")
            filepath = "{}/{}_fabs.csv".format(self.temporary_dir, generate_tas_rendering_label(tas))
            dump_to_csv(filepath, assistance_results)

            if self.verbose and (len(tas_dict_list) - i) % 10 == 0:
                logger.info("{:2>} TAS left".format(len(tas_dict_list) - i))

            if self.verbose:
                logger.info(self.contract_columns)
                logger.info(self.assistance_columns)

        logger.info("Completed fetching the data: {}s".format(perf_counter() - db_start))

    def single_tas_query(self, tas_dict, transaction_type="contract"):
        single_db_query = perf_counter()
        if transaction_type == "contract":
            sql_string = CONTRACT_SQL
        else:
            sql_string = ASSISTANCE_SQL
        formatted_dict = {k: "= '{}'".format(v) if v else "IS NULL" for k, v in tas_dict.items()}
        sql_string = sql_string.format(**formatted_dict)

        results = []
        with connection.cursor() as cursor:
            cursor.execute(sql_string)
            if transaction_type == "contract" and not self.contract_columns:
                self.contract_columns = [col[0] for col in cursor.description]
            elif transaction_type == "assistance" and not self.assistance_columns:
                self.assistance_columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
        if self.verbose:
            logger.info(json.dumps(tas_dict))
            logger.info(
                "DB query for {}s using above TAS took {}s and returned {} rows".format(
                    transaction_type, perf_counter() - single_db_query, len(results)
                )
            )
        return results

    def assemble_csv_files(self):
        logger.info("Reading CSVs")
        start_pandas = perf_counter()

        files = glob.glob(self.temporary_dir + "/*_fpds.csv")
        df = pd.concat([pd.read_csv(f, dtype=str, header=None, names=self.contract_columns) for f in files])
        self.write_pandas(df, "fpds")

        files = glob.glob(self.temporary_dir + "/*_fabs.csv")
        df = pd.concat([pd.read_csv(f, dtype=str, header=None, names=self.assistance_columns) for f in files])
        self.write_pandas(df, "fabs")

        logger.info("pandas took {}s".format(perf_counter() - start_pandas))

    def write_pandas(self, df, award_type):
        df = df.replace({np.nan: None})
        df = df.replace({"NaN": None})
        df["submission_period"] = pd.to_datetime(df["submission_period"])
        df["_fyq"] = df.apply(lambda x: generate_fiscal_year_and_quarter(x["submission_period"]), axis=1)

        logger.info("prepped pandas dataframe for all {} records".format(award_type))

        for quarter in pd.unique(df["_fyq"]):
            filepath = "{}/{}_{}.csv".format(self.destination, quarter, award_type)
            temp_df = df.loc[df["_fyq"] == quarter]
            del temp_df["_fyq"]
            temp_df.to_csv(path_or_buf=filepath, index=False)

    def cleanup_files(self):
        csv_files = glob.glob(self.destination + "/*.csv")
        zip_filepath = "{}_{}.zip".format(
            self.destination, datetime.utcnow().isoformat(sep="_", timespec="minutes")
        )
        zip_filepath.replace(":", "_")

        zipped_csvs = zipfile.ZipFile(zip_filepath, "a", compression=zipfile.ZIP_DEFLATED, allowZip64=True)
        for csv_file in csv_files:
            zipped_csvs.write(csv_file, os.path.basename(csv_file))

        if not self.keep_files:
            shutil.rmtree(self.temporary_dir)
            shutil.rmtree(self.destination)


ASSISTANCE_SQL = """
SELECT
    "financial_accounts_by_awards"."reporting_period_end" AS "submission_period",
    "treasury_appropriation_account"."allocation_transfer_agency_id" AS "allocation_transfer_agency_identifier",
    "treasury_appropriation_account"."agency_id" AS "agency_identifer",
    "treasury_appropriation_account"."beginning_period_of_availability" AS "beginning_period_of_availability",
    "treasury_appropriation_account"."ending_period_of_availability" AS "ending_period_of_availability",
    "treasury_appropriation_account"."availability_type_code" AS "availability_type_code",
    "treasury_appropriation_account"."main_account_code" AS "main_account_code",
    "treasury_appropriation_account"."sub_account_code" AS "sub_account_code",
    CONCAT("treasury_appropriation_account"."agency_id",
        CONCAT('-',
            CONCAT(CASE
                WHEN "treasury_appropriation_account"."availability_type_code" = 'X' THEN 'X'
                ELSE CONCAT("treasury_appropriation_account"."beginning_period_of_availability", CONCAT('/', "treasury_appropriation_account"."ending_period_of_availability"))
                END,
                CONCAT('-',
                    CONCAT("treasury_appropriation_account"."main_account_code",
                        CONCAT('-', "treasury_appropriation_account"."sub_account_code")
                    )
                )
            )
        )
    )
    AS "treasury_account_symbol",
    (SELECT U0."name" FROM "toptier_agency" U0 WHERE U0."cgac_code" = ( "treasury_appropriation_account"."agency_id") LIMIT 1)
      AS "agency_name",
    (SELECT U0. "name" FROM "toptier_agency" U0 WHERE U0."cgac_code" = ( "treasury_appropriation_account"."allocation_transfer_agency_id") LIMIT 1)
      AS "allocation_transfer_agency_name",
    "treasury_appropriation_account"."budget_function_title" AS "budget_function",
    "treasury_appropriation_account"."budget_subfunction_title" AS "budget_subfunction",
    CONCAT ("federal_account"."agency_identifier", CONCAT ('-', "federal_account"."main_account_code"))
      AS "federal_account_symbol",
    "federal_account"."account_title" AS "federal_account_name",
    "ref_program_activity"."program_activity_code" AS "program_activity_code",
    "ref_program_activity"."program_activity_name" AS "program_activity_name",
    "object_class"."object_class" AS "object_class_code",
    "object_class"."object_class_name" AS "object_class_name",
    "object_class"."direct_reimbursable" AS "direct_or_reimbursable_funding_source",
    "financial_accounts_by_awards"."piid" AS "piid",
    "financial_accounts_by_awards"."parent_award_id" AS "parent_award_piid",
    "financial_accounts_by_awards"."fain" AS "fain",
    "financial_accounts_by_awards"."uri" AS "uri",
    "financial_accounts_by_awards"."transaction_obligated_amount",

    "awards"."total_obligation" AS "obligated_amount", -- awards.total_obligation AS total_dollars_obligated,
    -- "awards"."potential_total_value_of_award",
    "awards"."total_funding_amount",
    "awards"."total_subsidy_cost",
    "awards"."total_loan_value",
    "awards"."period_of_performance_start_date",
    "awards"."period_of_performance_current_end_date",
    "subtier_agency"."subtier_code" AS "awarding_subagency_code",
    "subtier_agency"."name" AS "awarding_subagency_name",
    "awards"."type" AS "award_type_code",
    "awards"."type_description" AS "award_type",

    "transaction_fabs"."original_loan_subsidy_cost" AS "original_subsidy_cost",
    "transaction_fabs"."sai_number" AS "sai_number",
    "transaction_fabs"."non_federal_funding_amount" AS "non_federal_funding_amount",
    "transaction_fabs"."face_value_loan_guarantee" AS "face_value_of_loan",
    "transaction_fabs"."awarding_agency_code" AS "awarding_agency_code",
    "transaction_fabs"."awarding_agency_name" AS "awarding_agency_name",
    "transaction_fabs"."awarding_sub_tier_agency_c" AS "awarding_sub_agency_code",
    "transaction_fabs"."awarding_sub_tier_agency_n" AS "awarding_sub_agency_name",
    "transaction_fabs"."awarding_office_code" AS "awarding_office_code",
    "transaction_fabs"."awarding_office_name" AS "awarding_office_name",
    "transaction_fabs"."funding_agency_code" AS "funding_agency_code",
    "transaction_fabs"."funding_agency_name" AS "funding_agency_name",
    "transaction_fabs"."funding_sub_tier_agency_co" AS "funding_sub_agency_code",
    "transaction_fabs"."funding_sub_tier_agency_na" AS "funding_sub_agency_name",
    "transaction_fabs"."funding_office_code" AS "funding_office_code",
    "transaction_fabs"."funding_office_name" AS "funding_office_name",
    "transaction_fabs"."awardee_or_recipient_uniqu" AS "recipient_duns",
    "transaction_fabs"."awardee_or_recipient_legal" AS "recipient_name",
    "transaction_fabs"."ultimate_parent_unique_ide" AS "recipient_parent_duns",
    "transaction_fabs"."ultimate_parent_legal_enti" AS "recipient_parent_name",
    "transaction_fabs"."legal_entity_country_code" AS "recipient_country_code",
    "transaction_fabs"."legal_entity_country_name" AS "recipient_country_name",
    "transaction_fabs"."legal_entity_address_line1" AS "recipient_address_line_1",
    "transaction_fabs"."legal_entity_address_line2" AS "recipient_address_line_2",
    "transaction_fabs"."legal_entity_city_code" AS "recipient_city_code",
    "transaction_fabs"."legal_entity_city_name" AS "recipient_city_name",
    "transaction_fabs"."legal_entity_county_code" AS "recipient_county_code",
    "transaction_fabs"."legal_entity_county_name" AS "recipient_county_name",
    "transaction_fabs"."legal_entity_state_code" AS "recipient_state_code",
    "transaction_fabs"."legal_entity_state_name" AS "recipient_state_name",
    "transaction_fabs"."legal_entity_zip5" AS "recipient_zip_code",
    "transaction_fabs"."legal_entity_zip_last4" AS "recipient_zip_last_4_code",
    "transaction_fabs"."legal_entity_congressional" AS "recipient_congressional_district",
    "transaction_fabs"."legal_entity_foreign_city" AS "recipient_foreign_city_name",
    "transaction_fabs"."legal_entity_foreign_provi" AS "recipient_foreign_province_name",
    "transaction_fabs"."legal_entity_foreign_posta" AS "recipient_foreign_postal_code",
    "transaction_fabs"."place_of_perform_country_c" AS "primary_place_of_performance_country_code",
    "transaction_fabs"."place_of_perform_country_n" AS "primary_place_of_performance_country_name",
    "transaction_fabs"."place_of_performance_code" AS "primary_place_of_performance_code",
    "transaction_fabs"."place_of_performance_city" AS "primary_place_of_performance_city_name",
    "transaction_fabs"."place_of_perform_county_co" AS "primary_place_of_performance_county_code",
    "transaction_fabs"."place_of_perform_county_na" AS "primary_place_of_performance_county_name",
    "transaction_fabs"."place_of_perform_state_nam" AS "primary_place_of_performance_state_name",
    "transaction_fabs"."place_of_performance_zip4a" AS "primary_place_of_performance_zip_4",
    "transaction_fabs"."place_of_performance_congr" AS "primary_place_of_performance_congressional_district",
    "transaction_fabs"."place_of_performance_forei" AS "primary_place_of_performance_foreign_location",
    "transaction_fabs"."cfda_number" AS "cfda_number",
    "transaction_fabs"."cfda_title" AS "cfda_title",
    "transaction_fabs"."assistance_type" AS "assistance_type_code",
    "transaction_fabs"."assistance_type_desc" AS "assistance_type_description",
    "transaction_fabs"."award_description" AS "award_description",
    "transaction_fabs"."business_funds_indicator" AS "business_funds_indicator_code",
    "transaction_fabs"."business_funds_ind_desc" AS "business_funds_indicator_description",
    "transaction_fabs"."business_types" AS "business_types_code",
    "transaction_fabs"."business_types_desc" AS "business_types_description",
    "transaction_fabs"."record_type" AS "record_type_code",
    "transaction_fabs"."record_type_description" AS "record_type_description",
    "transaction_fabs"."modified_at" AS "last_modified_date"
    FROM
    "financial_accounts_by_awards"
    INNER JOIN
    "treasury_appropriation_account"
        ON (
            "financial_accounts_by_awards"."treasury_account_id" = "treasury_appropriation_account"."treasury_account_identifier"
       )
    LEFT OUTER JOIN
    "federal_account"
        ON (
            "treasury_appropriation_account"."federal_account_id" = "federal_account"."id"
       )
    LEFT OUTER JOIN
    "awards"
        ON (
            "financial_accounts_by_awards"."award_id" = "awards"."id"
       )
    INNER JOIN
    "transaction_fabs"
        ON (
            "awards"."latest_transaction_id" = "transaction_fabs"."transaction_id"
       )
    LEFT OUTER JOIN
    "ref_program_activity"
        ON (
            "financial_accounts_by_awards"."program_activity_id" = "ref_program_activity"."id"
       )
    LEFT OUTER JOIN
    "object_class"
        ON (
            "financial_accounts_by_awards"."object_class_id" = "object_class"."id"
       )
    LEFT OUTER JOIN
    "agency"
        ON (
            "awards"."awarding_agency_id" = "agency"."id"
       )
    LEFT OUTER JOIN
    "toptier_agency"
        ON (
            "agency"."toptier_agency_id" = "toptier_agency"."toptier_agency_id"
       )
    LEFT OUTER JOIN
    "subtier_agency"
        ON (
            "agency"."subtier_agency_id" = "subtier_agency"."subtier_agency_id"
       )
    WHERE
    (
        (
            treasury_appropriation_account.allocation_transfer_agency_id {ATA} AND
            treasury_appropriation_account.agency_id {AID} AND
            treasury_appropriation_account.availability_type_code {AvailType Code} AND
            treasury_appropriation_account.beginning_period_of_availability {BPOA} AND
            treasury_appropriation_account.ending_period_of_availability {EPOA} AND
            treasury_appropriation_account.main_account_code {Main} AND
            treasury_appropriation_account.sub_account_code {Sub}
        )
    )
;
"""


CONTRACT_SQL = """
SELECT
    "financial_accounts_by_awards"."reporting_period_end" AS "submission_period",
    "treasury_appropriation_account"."allocation_transfer_agency_id" AS "allocation_transfer_agency_identifier",
    "treasury_appropriation_account"."agency_id" AS "agency_identifer",
    "treasury_appropriation_account"."beginning_period_of_availability" AS "beginning_period_of_availability",
    "treasury_appropriation_account"."ending_period_of_availability" AS "ending_period_of_availability",
    "treasury_appropriation_account"."availability_type_code" AS "availability_type_code",
    "treasury_appropriation_account"."main_account_code" AS "main_account_code",
    "treasury_appropriation_account"."sub_account_code" AS "sub_account_code",
    CONCAT("treasury_appropriation_account"."agency_id",
        CONCAT('-',
            CONCAT(CASE
                WHEN "treasury_appropriation_account"."availability_type_code" = 'X' THEN 'X'
                ELSE CONCAT("treasury_appropriation_account"."beginning_period_of_availability", CONCAT('/', "treasury_appropriation_account"."ending_period_of_availability"))
                END,
                CONCAT('-',
                    CONCAT("treasury_appropriation_account"."main_account_code",
                        CONCAT('-', "treasury_appropriation_account"."sub_account_code")
                    )
                )
            )
        )
    )
    AS "treasury_account_symbol",
    (SELECT U0."name" FROM "toptier_agency" U0 WHERE U0."cgac_code" = ( "treasury_appropriation_account"."agency_id") LIMIT 1)
      AS "agency_name",
    (SELECT U0. "name" FROM "toptier_agency" U0 WHERE U0."cgac_code" = ( "treasury_appropriation_account"."allocation_transfer_agency_id") LIMIT 1)
      AS "allocation_transfer_agency_name",
    "treasury_appropriation_account"."budget_function_title" AS "budget_function",
    "treasury_appropriation_account"."budget_subfunction_title" AS "budget_subfunction",
    CONCAT ("federal_account"."agency_identifier", CONCAT ('-', "federal_account"."main_account_code"))
      AS "federal_account_symbol",
    "federal_account"."account_title" AS "federal_account_name",
    "ref_program_activity"."program_activity_code" AS "program_activity_code",
    "ref_program_activity"."program_activity_name" AS "program_activity_name",
    "object_class"."object_class" AS "object_class_code",
    "object_class"."object_class_name" AS "object_class_name",
    "object_class"."direct_reimbursable" AS "direct_or_reimbursable_funding_source",
    "financial_accounts_by_awards"."piid" AS "piid",
    "financial_accounts_by_awards"."parent_award_id" AS "parent_award_piid",
    "financial_accounts_by_awards"."fain" AS "fain",
    "financial_accounts_by_awards"."uri" AS "uri",
    "financial_accounts_by_awards"."transaction_obligated_amount",
    "awards"."total_obligation" AS "obligated_amount", -- awards.total_obligation AS total_dollars_obligated,
    "subtier_agency"."subtier_code" AS "awarding_subagency_code",
    "subtier_agency"."name" AS "awarding_subagency_name",

    "transaction_fpds"."referenced_idv_agency_iden" AS "parent_award_agency_id",
    "transaction_fpds"."referenced_idv_agency_desc" AS "parent_award_agency_name",
    "transaction_fpds"."current_total_value_award" AS "current_total_value_of_award",
    "transaction_fpds"."potential_total_value_awar" AS "potential_total_value_of_award",
    "transaction_fpds"."period_of_performance_star" AS "period_of_performance_start_date",
    "transaction_fpds"."period_of_performance_curr" AS "period_of_performance_current_end_date",
    "transaction_fpds"."period_of_perf_potential_e" AS "period_of_performance_potential_end_date",
    "transaction_fpds"."ordering_period_end_date" AS "ordering_period_end_date",
    "transaction_fpds"."awarding_agency_code" AS "awarding_agency_code",
    "transaction_fpds"."awarding_agency_name" AS "awarding_agency_name",
    "transaction_fpds"."awarding_sub_tier_agency_c" AS "awarding_sub_agency_code",
    "transaction_fpds"."awarding_sub_tier_agency_n" AS "awarding_sub_agency_name",
    "transaction_fpds"."awarding_office_code" AS "awarding_office_code",
    "transaction_fpds"."awarding_office_name" AS "awarding_office_name",
    "transaction_fpds"."funding_agency_code" AS "funding_agency_code",
    "transaction_fpds"."funding_agency_name" AS "funding_agency_name",
    "transaction_fpds"."funding_sub_tier_agency_co" AS "funding_sub_agency_code",
    "transaction_fpds"."funding_sub_tier_agency_na" AS "funding_sub_agency_name",
    "transaction_fpds"."funding_office_code" AS "funding_office_code",
    "transaction_fpds"."funding_office_name" AS "funding_office_name",
    "transaction_fpds"."foreign_funding" AS "foreign_funding",
    "transaction_fpds"."foreign_funding_desc" AS "foreign_funding_description",
    "transaction_fpds"."sam_exception" AS "sam_exception",
    "transaction_fpds"."sam_exception_description" AS "sam_exception_description",
    "transaction_fpds"."awardee_or_recipient_uniqu" AS "recipient_duns",
    "transaction_fpds"."awardee_or_recipient_legal" AS "recipient_name",
    "transaction_fpds"."vendor_doing_as_business_n" AS "recipient_doing_business_as_name",
    "transaction_fpds"."cage_code" AS "cage_code",
    "transaction_fpds"."ultimate_parent_legal_enti" AS "recipient_parent_name",
    "transaction_fpds"."ultimate_parent_unique_ide" AS "recipient_parent_duns",
    "transaction_fpds"."legal_entity_country_code" AS "recipient_country_code",
    "transaction_fpds"."legal_entity_country_name" AS "recipient_country_name",
    "transaction_fpds"."legal_entity_address_line1" AS "recipient_address_line_1",
    "transaction_fpds"."legal_entity_address_line2" AS "recipient_address_line_2",
    "transaction_fpds"."legal_entity_city_name" AS "recipient_city_name",
    "transaction_fpds"."legal_entity_state_code" AS "recipient_state_code",
    "transaction_fpds"."legal_entity_state_descrip" AS "recipient_state_name",
    "transaction_fpds"."legal_entity_zip4" AS "recipient_zip_4_code",
    "transaction_fpds"."legal_entity_congressional" AS "recipient_congressional_district",
    "transaction_fpds"."vendor_phone_number" AS "recipient_phone_number",
    "transaction_fpds"."vendor_fax_number" AS "recipient_fax_number",
    "transaction_fpds"."place_of_perform_country_c" AS "primary_place_of_performance_country_code",
    "transaction_fpds"."place_of_perf_country_desc" AS "primary_place_of_performance_country_name",
    "transaction_fpds"."place_of_perform_city_name" AS "primary_place_of_performance_city_name",
    "transaction_fpds"."place_of_perform_county_na" AS "primary_place_of_performance_county_name",
    "transaction_fpds"."place_of_performance_state" AS "primary_place_of_performance_state_code",
    "transaction_fpds"."place_of_perfor_state_desc" AS "primary_place_of_performance_state_name",
    "transaction_fpds"."place_of_performance_zip4a" AS "primary_place_of_performance_zip_4",
    "transaction_fpds"."place_of_performance_congr" AS "primary_place_of_performance_congressional_district",
    "transaction_fpds"."pulled_from" AS "award_or_idv_flag",
    "transaction_fpds"."contract_award_type" AS "award_type_code",
    "transaction_fpds"."contract_award_type_desc" AS "award_type",
    "transaction_fpds"."idv_type" AS "idv_type_code",
    "transaction_fpds"."idv_type_description" AS "idv_type",
    "transaction_fpds"."multiple_or_single_award_i" AS "multiple_or_single_award_idv_code",
    "transaction_fpds"."multiple_or_single_aw_desc" AS "multiple_or_single_award_idv",
    "transaction_fpds"."type_of_idc" AS "type_of_idc_code",
    "transaction_fpds"."type_of_idc_description" AS "type_of_idc",
    "transaction_fpds"."type_of_contract_pricing" AS "type_of_contract_pricing_code",
    "transaction_fpds"."type_of_contract_pric_desc" AS "type_of_contract_pricing",
    "transaction_fpds"."award_description" AS "award_description",
    "transaction_fpds"."solicitation_identifier" AS "solicitation_identifier",
    "transaction_fpds"."number_of_actions" AS "number_of_actions",
    "transaction_fpds"."product_or_service_code" AS "product_or_service_code",
    "transaction_fpds"."product_or_service_co_desc" AS "product_or_service_code_description",
    "transaction_fpds"."contract_bundling" AS "contract_bundling_code",
    "transaction_fpds"."contract_bundling_descrip" AS "contract_bundling",
    "transaction_fpds"."dod_claimant_program_code" AS "dod_claimant_program_code",
    "transaction_fpds"."dod_claimant_prog_cod_desc" AS "dod_claimant_program_description",
    "transaction_fpds"."naics" AS "naics_code",
    "transaction_fpds"."naics_description" AS "naics_description",
    "transaction_fpds"."recovered_materials_sustai" AS "recovered_materials_sustainability_code",
    "transaction_fpds"."recovered_materials_s_desc" AS "recovered_materials_sustainability",
    "transaction_fpds"."domestic_or_foreign_entity" AS "domestic_or_foreign_entity_code",
    "transaction_fpds"."domestic_or_foreign_e_desc" AS "domestic_or_foreign_entity",
    "transaction_fpds"."program_system_or_equipmen" AS "dod_acquisition_program_code",
    "transaction_fpds"."program_system_or_equ_desc" AS "dod_acquisition_program_description",
    "transaction_fpds"."information_technology_com" AS "information_technology_commercial_item_category_code",
    "transaction_fpds"."information_technolog_desc" AS "information_technology_commercial_item_category",
    "transaction_fpds"."epa_designated_product" AS "epa_designated_product_code",
    "transaction_fpds"."epa_designated_produc_desc" AS "epa_designated_product",
    "transaction_fpds"."country_of_product_or_serv" AS "country_of_product_or_service_origin_code",
    "transaction_fpds"."country_of_product_or_desc" AS "country_of_product_or_service_origin",
    "transaction_fpds"."place_of_manufacture" AS "place_of_manufacture_code",
    "transaction_fpds"."place_of_manufacture_desc" AS "place_of_manufacture",
    "transaction_fpds"."subcontracting_plan" AS "subcontracting_plan_code",
    "transaction_fpds"."subcontracting_plan_desc" AS "subcontracting_plan",
    "transaction_fpds"."extent_competed" AS "extent_competed_code",
    "transaction_fpds"."extent_compete_description" AS "extent_competed",
    "transaction_fpds"."solicitation_procedures" AS "solicitation_procedures_code",
    "transaction_fpds"."solicitation_procedur_desc" AS "solicitation_procedures",
    "transaction_fpds"."type_set_aside" AS "type_of_set_aside_code",
    "transaction_fpds"."type_set_aside_description" AS "type_of_set_aside",
    "transaction_fpds"."evaluated_preference" AS "evaluated_preference_code",
    "transaction_fpds"."evaluated_preference_desc" AS "evaluated_preference",
    "transaction_fpds"."research" AS "research_code",
    "transaction_fpds"."research_description" AS "research",
    "transaction_fpds"."fair_opportunity_limited_s" AS "fair_opportunity_limited_sources_code",
    "transaction_fpds"."fair_opportunity_limi_desc" AS "fair_opportunity_limited_sources",
    "transaction_fpds"."other_than_full_and_open_c" AS "other_than_full_and_open_competition_code",
    "transaction_fpds"."other_than_full_and_o_desc" AS "other_than_full_and_open_competition",
    "transaction_fpds"."number_of_offers_received" AS "number_of_offers_received",
    "transaction_fpds"."commercial_item_acquisitio" AS "commercial_item_acquisition_procedures_code",
    "transaction_fpds"."commercial_item_acqui_desc" AS "commercial_item_acquisition_procedures",
    "transaction_fpds"."small_business_competitive" AS "small_business_competitiveness_demonstration_program",
    "transaction_fpds"."commercial_item_test_progr" AS "commercial_item_test_program_code",
    "transaction_fpds"."commercial_item_test_desc" AS "commercial_item_test_program",
    "transaction_fpds"."a_76_fair_act_action" AS "a76_fair_act_action_code",
    "transaction_fpds"."a_76_fair_act_action_desc" AS "a76_fair_act_action",
    "transaction_fpds"."fed_biz_opps" AS "fed_biz_opps_code",
    "transaction_fpds"."fed_biz_opps_description" AS "fed_biz_opps",
    "transaction_fpds"."local_area_set_aside" AS "local_area_set_aside_code",
    "transaction_fpds"."local_area_set_aside_desc" AS "local_area_set_aside",
    "transaction_fpds"."clinger_cohen_act_planning" AS "clinger_cohen_act_planning_code",
    "transaction_fpds"."clinger_cohen_act_pla_desc" AS "clinger_cohen_act_planning",
    "transaction_fpds"."materials_supplies_article" AS "materials_supplies_articles_equipment_code",
    "transaction_fpds"."materials_supplies_descrip" AS "materials_supplies_articles_equipment",
    "transaction_fpds"."labor_standards" AS "labor_standards_code",
    "transaction_fpds"."labor_standards_descrip" AS "labor_standards",
    "transaction_fpds"."construction_wage_rate_req" AS "construction_wage_rate_requirements_code",
    "transaction_fpds"."construction_wage_rat_desc" AS "construction_wage_rate_requirements",
    "transaction_fpds"."interagency_contracting_au" AS "interagency_contracting_authority_code",
    "transaction_fpds"."interagency_contract_desc" AS "interagency_contracting_authority",
    "transaction_fpds"."other_statutory_authority" AS "other_statutory_authority",
    "transaction_fpds"."program_acronym" AS "program_acronym",
    "transaction_fpds"."referenced_idv_type" AS "parent_award_type_code",
    "transaction_fpds"."referenced_idv_type_desc" AS "parent_award_type",
    "transaction_fpds"."referenced_mult_or_single" AS "parent_award_single_or_multiple_code",
    "transaction_fpds"."referenced_mult_or_si_desc" AS "parent_award_single_or_multiple",
    "transaction_fpds"."major_program" AS "major_program",
    "transaction_fpds"."national_interest_action" AS "national_interest_action_code",
    "transaction_fpds"."national_interest_desc" AS "national_interest_action",
    "transaction_fpds"."cost_or_pricing_data" AS "cost_or_pricing_data_code",
    "transaction_fpds"."cost_or_pricing_data_desc" AS "cost_or_pricing_data",
    "transaction_fpds"."cost_accounting_standards" AS "cost_accounting_standards_clause_code",
    "transaction_fpds"."cost_accounting_stand_desc" AS "cost_accounting_standards_clause",
    "transaction_fpds"."government_furnished_prope" AS "gfe_gfp_code",
    "transaction_fpds"."government_furnished_prope" AS "gfe_gfp",
    "transaction_fpds"."sea_transportation" AS "sea_transportation_code",
    "transaction_fpds"."sea_transportation_desc" AS "sea_transportation",
    "transaction_fpds"."consolidated_contract" AS "consolidated_contract_code",
    "transaction_fpds"."consolidated_contract_desc" AS "consolidated_contract",
    "transaction_fpds"."performance_based_service" AS "performance_based_service_acquisition_code",
    "transaction_fpds"."performance_based_se_desc" AS "performance_based_service_acquisition",
    "transaction_fpds"."multi_year_contract" AS "multi_year_contract_code",
    "transaction_fpds"."multi_year_contract_desc" AS "multi_year_contract",
    "transaction_fpds"."contract_financing" AS "contract_financing_code",
    "transaction_fpds"."contract_financing_descrip" AS "contract_financing",
    "transaction_fpds"."purchase_card_as_payment_m" AS "purchase_card_as_payment_method_code",
    "transaction_fpds"."purchase_card_as_paym_desc" AS "purchase_card_as_payment_method",
    "transaction_fpds"."contingency_humanitarian_o" AS "contingency_humanitarian_or_peacekeeping_operation_code",
    "transaction_fpds"."contingency_humanitar_desc" AS "contingency_humanitarian_or_peacekeeping_operation",
    "transaction_fpds"."alaskan_native_owned_corpo" AS "alaskan_native_owned_corporation_or_firm",
    "transaction_fpds"."american_indian_owned_busi" AS "american_indian_owned_business",
    "transaction_fpds"."indian_tribe_federally_rec" AS "indian_tribe_federally_recognized",
    "transaction_fpds"."native_hawaiian_owned_busi" AS "native_hawaiian_owned_business",
    "transaction_fpds"."tribally_owned_business" AS "tribally_owned_business",
    "transaction_fpds"."veteran_owned_business" AS "veteran_owned_business",
    "transaction_fpds"."service_disabled_veteran_o" AS "service_disabled_veteran_owned_business",
    "transaction_fpds"."woman_owned_business" AS "woman_owned_business",
    "transaction_fpds"."women_owned_small_business" AS "women_owned_small_business",
    "transaction_fpds"."economically_disadvantaged" AS "economically_disadvantaged_women_owned_small_business",
    "transaction_fpds"."joint_venture_women_owned" AS "joint_venture_women_owned_small_business",
    "transaction_fpds"."joint_venture_economically" AS "joint_venture_economic_disadvantaged_women_owned_small_bus",
    "transaction_fpds"."minority_owned_business" AS "minority_owned_business",
    "transaction_fpds"."subcontinent_asian_asian_i" AS "subcontinent_asian_asian_indian_american_owned_business",
    "transaction_fpds"."asian_pacific_american_own" AS "asian_pacific_american_owned_business",
    "transaction_fpds"."black_american_owned_busin" AS "black_american_owned_business",
    "transaction_fpds"."hispanic_american_owned_bu" AS "hispanic_american_owned_business",
    "transaction_fpds"."native_american_owned_busi" AS "native_american_owned_business",
    "transaction_fpds"."other_minority_owned_busin" AS "other_minority_owned_business",
    "transaction_fpds"."contracting_officers_desc" AS "contracting_officers_determination_of_business_size",
    "transaction_fpds"."contracting_officers_deter" AS "contracting_officers_determination_of_business_size_code",
    "transaction_fpds"."emerging_small_business" AS "emerging_small_business",
    "transaction_fpds"."community_developed_corpor" AS "community_developed_corporation_owned_firm",
    "transaction_fpds"."labor_surplus_area_firm" AS "labor_surplus_area_firm",
    "transaction_fpds"."us_federal_government" AS "us_federal_government",
    "transaction_fpds"."federally_funded_research" AS "federally_funded_research_and_development_corp",
    "transaction_fpds"."federal_agency" AS "federal_agency",
    "transaction_fpds"."us_state_government" AS "us_state_government",
    "transaction_fpds"."us_local_government" AS "us_local_government",
    "transaction_fpds"."city_local_government" AS "city_local_government",
    "transaction_fpds"."county_local_government" AS "county_local_government",
    "transaction_fpds"."inter_municipal_local_gove" AS "inter_municipal_local_government",
    "transaction_fpds"."local_government_owned" AS "local_government_owned",
    "transaction_fpds"."municipality_local_governm" AS "municipality_local_government",
    "transaction_fpds"."school_district_local_gove" AS "school_district_local_government",
    "transaction_fpds"."township_local_government" AS "township_local_government",
    "transaction_fpds"."us_tribal_government" AS "us_tribal_government",
    "transaction_fpds"."foreign_government" AS "foreign_government",
    "transaction_fpds"."organizational_type" AS "organizational_type",
    "transaction_fpds"."corporate_entity_not_tax_e" AS "corporate_entity_not_tax_exempt",
    "transaction_fpds"."corporate_entity_tax_exemp" AS "corporate_entity_tax_exempt",
    "transaction_fpds"."partnership_or_limited_lia" AS "partnership_or_limited_liability_partnership",
    "transaction_fpds"."sole_proprietorship" AS "sole_proprietorship",
    "transaction_fpds"."small_agricultural_coopera" AS "small_agricultural_cooperative",
    "transaction_fpds"."international_organization" AS "international_organization",
    "transaction_fpds"."us_government_entity" AS "us_government_entity",
    "transaction_fpds"."community_development_corp" AS "community_development_corporation",
    "transaction_fpds"."domestic_shelter" AS "domestic_shelter",
    "transaction_fpds"."educational_institution" AS "educational_institution",
    "transaction_fpds"."foundation" AS "foundation",
    "transaction_fpds"."hospital_flag" AS "hospital_flag",
    "transaction_fpds"."manufacturer_of_goods" AS "manufacturer_of_goods",
    "transaction_fpds"."veterinary_hospital" AS "veterinary_hospital",
    "transaction_fpds"."hispanic_servicing_institu" AS "hispanic_servicing_institution",
    "transaction_fpds"."contracts" AS "receives_contracts",
    "transaction_fpds"."grants" AS "receives_grants",
    "transaction_fpds"."receives_contracts_and_gra" AS "receives_contracts_and_grants",
    "transaction_fpds"."airport_authority" AS "airport_authority",
    "transaction_fpds"."council_of_governments" AS "council_of_governments",
    "transaction_fpds"."housing_authorities_public" AS "housing_authorities_public_tribal",
    "transaction_fpds"."interstate_entity" AS "interstate_entity",
    "transaction_fpds"."planning_commission" AS "planning_commission",
    "transaction_fpds"."port_authority" AS "port_authority",
    "transaction_fpds"."transit_authority" AS "transit_authority",
    "transaction_fpds"."subchapter_s_corporation" AS "subchapter_scorporation",
    "transaction_fpds"."limited_liability_corporat" AS "limited_liability_corporation",
    "transaction_fpds"."foreign_owned_and_located" AS "foreign_owned_and_located",
    "transaction_fpds"."for_profit_organization" AS "for_profit_organization",
    "transaction_fpds"."nonprofit_organization" AS "nonprofit_organization",
    "transaction_fpds"."other_not_for_profit_organ" AS "other_not_for_profit_organization",
    "transaction_fpds"."the_ability_one_program" AS "the_ability_one_program",
    "transaction_fpds"."number_of_employees" AS "number_of_employees",
    "transaction_fpds"."annual_revenue" AS "annual_revenue",
    "transaction_fpds"."private_university_or_coll" AS "private_university_or_college",
    "transaction_fpds"."state_controlled_instituti" AS "state_controlled_institution_of_higher_learning",
    "transaction_fpds"."c1862_land_grant_college" AS "1862_land_grant_college",
    "transaction_fpds"."c1890_land_grant_college" AS "1890_land_grant_college",
    "transaction_fpds"."c1994_land_grant_college" AS "1994_land_grant_college",
    "transaction_fpds"."minority_institution" AS "minority_institution",
    "transaction_fpds"."historically_black_college" AS "historically_black_college",
    "transaction_fpds"."tribal_college" AS "tribal_college",
    "transaction_fpds"."alaskan_native_servicing_i" AS "alaskan_native_servicing_institution",
    "transaction_fpds"."native_hawaiian_servicing" AS "native_hawaiian_servicing_institution",
    "transaction_fpds"."school_of_forestry" AS "school_of_forestry",
    "transaction_fpds"."veterinary_college" AS "veterinary_college",
    "transaction_fpds"."dot_certified_disadvantage" AS "dot_certified_disadvantage",
    "transaction_fpds"."self_certified_small_disad" AS "self_certified_small_disadvantaged_business",
    "transaction_fpds"."small_disadvantaged_busine" AS "small_disadvantaged_business",
    "transaction_fpds"."c8a_program_participant" AS "c8a_program_participant",
    "transaction_fpds"."historically_underutilized" AS "historically_underutilized_business_zone_hubzone_firm",
    "transaction_fpds"."sba_certified_8_a_joint_ve" AS "sba_certified_8a_joint_venture",
    "transaction_fpds"."last_modified" AS "last_modified_date"
    FROM
    "financial_accounts_by_awards"
    INNER JOIN
    "treasury_appropriation_account"
        ON (
            "financial_accounts_by_awards"."treasury_account_id" = "treasury_appropriation_account"."treasury_account_identifier"
       )
    LEFT OUTER JOIN
    "federal_account"
        ON (
            "treasury_appropriation_account"."federal_account_id" = "federal_account"."id"
       )
    LEFT OUTER JOIN
    "awards"
        ON (
            "financial_accounts_by_awards"."award_id" = "awards"."id"
       )
    INNER JOIN
    "transaction_fpds"
        ON (
            "awards"."latest_transaction_id" = "transaction_fpds"."transaction_id"
       )
    LEFT OUTER JOIN
    "ref_program_activity"
        ON (
            "financial_accounts_by_awards"."program_activity_id" = "ref_program_activity"."id"
       )
    LEFT OUTER JOIN
    "object_class"
        ON (
            "financial_accounts_by_awards"."object_class_id" = "object_class"."id"
       )
    LEFT OUTER JOIN
    "agency"
        ON (
            "awards"."awarding_agency_id" = "agency"."id"
       )
    LEFT OUTER JOIN
    "toptier_agency"
        ON (
            "agency"."toptier_agency_id" = "toptier_agency"."toptier_agency_id"
       )
    LEFT OUTER JOIN
    "subtier_agency"
        ON (
            "agency"."subtier_agency_id" = "subtier_agency"."subtier_agency_id"
       )
    WHERE
    (
        (
            treasury_appropriation_account.allocation_transfer_agency_id {ATA} AND
            treasury_appropriation_account.agency_id {AID} AND
            treasury_appropriation_account.availability_type_code {AvailType Code} AND
            treasury_appropriation_account.beginning_period_of_availability {BPOA} AND
            treasury_appropriation_account.ending_period_of_availability {EPOA} AND
            treasury_appropriation_account.main_account_code {Main} AND
            treasury_appropriation_account.sub_account_code {Sub}
        )
    )
;"""
